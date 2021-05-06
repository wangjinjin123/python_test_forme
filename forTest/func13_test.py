"""
python外部数据源文件处理
Yaml：
是一种可读性高，用来表达数序列化的格式，常常作为配置文件使用
Json：
是一个轻量级的数据交换语言，该语言以易于让人阅读的文字为基础，用来传输由属性值或者需理性的值组成的数据对象
Excel：
有直观的界面、出色的计算功能和图标工具是一款电子制表软件
"""


from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
dest_filename = 'empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")

ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
     for col in range(27, 54):
         _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)

ws4 = wb.create_sheet(title="my_sheet")
for i in range(1,31):
    ws4.cell(column=1,row=i).value="test"
wb.save(filename = dest_filename)



#读数据
import yaml
# from openpyxl import load_workbook
# wb = load_workbook(filename = 'empty_book.xlsx')
# sheet_ranges = wb['range names']
# print(sheet_ranges['D18'].value)
#
# for i in range(1,31):
#     print(sheet_ranges.cell(column=1, row=i).value)



#json的读写



"""
Yaml的读写  推荐
PyYAML
  yaml.load   yaml格式转成其他格式
  yaml.dump   其他格式转yaml格式
"""

#loading YAML
# print(yaml.load("""
#  - Hesperiidae
#  - Papilionidae
#  - Apatelodidae
#  - Epiplemidae
#  """, Loader=yaml.FullLoader))


print(yaml.load(open("../data/yaml_test.yml"), Loader=yaml.FullLoader))


#dumping YAML
print(yaml.dump([['Hesperiidae', 'Papilionidae', 'Apatelodidae', 'Epiplemidae', {'a': 1}]]))
with open("yaml1_test.yml","w") as f:
    yaml.dump(data={"a": [1,2]}, stream=f)